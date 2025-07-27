from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from datetime import timedelta
import math
import re
import matplotlib.pyplot as plt
import io
import base64
import matplotlib.ticker as ticker
import numpy as np
import math
from scipy.interpolate import CubicSpline , interp1d , Akima1DInterpolator
from scipy.optimize import minimize_scalar
from io import BytesIO
from scipy.interpolate import make_interp_spline
from odoo.exceptions import UserError
import openpyxl


class EnviromentTemprature(models.Model):
    _name = 'enviroment.temprature'  
    _inherit = "lerm.eln" 

    _rec_name = "name"

    name = fields.Char("Name",default="Temperature Monitoring Of  Concreate")
    parameter_id = fields.Many2one('eln.parameters.result',string="Parameter")
    sample_parameters = fields.Many2many('lerm.parameter.master',string="Parameters",compute="_compute_sample_parameters",store=True)
    eln_ref = fields.Many2one('lerm.eln',string="Eln")
    grade = fields.Many2one('lerm.grade.line',string="Grade",compute="_compute_grade_id",store=True)

    @api.depends('eln_ref')
    def _compute_grade_id(self):
        if self.eln_ref:
            self.grade = self.eln_ref.grade_id.id

    # lab_name = fields.Selection([
    #         ('wadala', 'WADALA'),
    #         ('taloja', 'TALOJA')], string="Lab Name")
    
    # child_lines_temprature = fields.One2many('enviroment.tempraturee.line','parent_id',string="Parameter", copy=True)

    

    # peak_max_temp = fields.Char(
    #     string="Peak Concrete Temperature ",  compute="_compute_peak_max_middle1", store=True
    #   )

    # peak_max_temp_middle1_observed = fields.Char(
    #     string="Peak Concrete Temperature Observed At ", compute="_compute_peak_max_middle1", store=True
    #   )



    # peak_max_temp_diff = fields.Char(
    #     string="Maximum Thermal Gradient ",  compute="_compute_peak_max_temp_diff", store=True
    #   )

    # peak_max_temp_diff_observed = fields.Char(
    #     string="Maximum Thermal Gradient Observed At",  compute="_compute_peak_max_temp_diff", store=True
    #   )

    # peak_max_temp_ambient = fields.Char(
    #     string="Peak Ambient Temperature  ",digits=(12,1),  compute="_compute_peak_max_ambient", store=True
    #   )

    # peak_max_temp_ambient_observed = fields.Char(
    #     string="Peak Ambient Temperature Observed At " ,compute="_compute_peak_max_ambient",store=True
    #   )



    # @api.depends('child_lines_temprature.temp_diff')
    # def _compute_peak_max_temp_diff(self):
    #     for record in self:
    #         max_temp_diff = max(
    #             record.child_lines_temprature.mapped('temp_diff') or [0]
    #         )  # Use `or [0]` to handle cases with no child lines
    #         record.peak_max_temp_diff = max_temp_diff


    # @api.depends('child_lines_temprature.ambient')
    # def _compute_peak_max_ambient(self):
    #     for record in self:
    #         if record.child_lines_temprature:
    #             max_line = max(record.child_lines_temprature, key=lambda line: line.ambient)
    #             max_ambient = max_line.ambient
    #             measurement_time = max_line.measurement_time  # Assuming this field exists in `child_lines_temprature`
    #             day = max_line.day  # Assuming this field exists in `child_lines_temprature`
                
    #             if measurement_time and day:
    #                 formatted_time = measurement_time
    #                 formatted_day = day
    #                 record.peak_max_temp_ambient_observed = f"At {formatted_time} on {formatted_day}"
    #             else:
    #                 record.peak_max_temp_ambient_observed = "No measurement time available"
                
    #             # record.peak_max_temp_ambient = max_ambient
    #             record.peak_max_temp_ambient = f"{max_ambient} °C"
    #         else:
    #             # Default values if no child lines exist
    #             record.peak_max_temp_ambient = 0.0
    #             record.peak_max_temp_ambient_observed = "No data"



    # @api.depends('child_lines_temprature.middle1')
    # def _compute_peak_max_middle1(self):
    #     for record in self:
    #         if record.child_lines_temprature:
    #             # Find the maximum value of middle1
    #             max_middle1 = max(line.middle1 for line in record.child_lines_temprature)
                
    #             # Filter lines with the maximum value of middle1
    #             max_lines = [line for line in record.child_lines_temprature if line.middle1 == max_middle1]
                
    #             # Extract times and days from the matching lines
    #             if max_lines:
    #                 measurement_times = []
    #                 day = max_lines[0].day  # Assuming all lines with max middle1 have the same day
                    
    #                 for max_line in max_lines:
    #                     measurement_time = max_line.measurement_time
    #                     if measurement_time:
    #                         measurement_times.append(measurement_time)
                    
    #                 # Format observed data
    #                 if len(measurement_times) == 1:
    #                     record.peak_max_temp_middle1_observed = f"At {measurement_times[0]} on {day}"
    #                 elif len(measurement_times) > 1:
    #                     record.peak_max_temp_middle1_observed = (
    #                         f"Between {measurement_times[0]} - {measurement_times[-1]} on {day}"
    #                     )
    #                 else:
    #                     record.peak_max_temp_middle1_observed = "No measurement time available"
                    
    #                 # Set the peak max temperature with °C
    #                 record.peak_max_temp = f"{max_middle1:.1f} °C"
    #         else:
    #             # Default values when no child lines exist
    #             record.peak_max_temp = "0.0 °C"
    #             record.peak_max_temp_middle1_observed = "No data"


    # @api.depends('child_lines_temprature.temp_diff')
    # def _compute_peak_max_temp_diff(self):
    #     for record in self:
    #         if record.child_lines_temprature:
    #             # Find the maximum value of temp_diff
    #             max_temp_diff = max(line.temp_diff for line in record.child_lines_temprature)
                
    #             # Filter lines with the maximum value of temp_diff
    #             max_lines = [line for line in record.child_lines_temprature if line.temp_diff == max_temp_diff]
                
    #             # Extract times and days from the matching lines
    #             if max_lines:
    #                 measurement_times = []
    #                 day = max_lines[0].day  # Assuming all lines with max temp_diff have the same day
                    
    #                 for max_line in max_lines:
    #                     measurement_time = max_line.measurement_time
    #                     if measurement_time:
    #                         measurement_times.append(measurement_time)
                    
    #                 # Format observed data
    #                 if len(measurement_times) == 1:
    #                     record.peak_max_temp_diff_observed = f"At {measurement_times[0]} on {day}"
    #                 elif len(measurement_times) > 1:
    #                     record.peak_max_temp_diff_observed = (
    #                         f"Between {measurement_times[0]} - {measurement_times[-1]} on {day}"
    #                     )
    #                 else:
    #                     record.peak_max_temp_diff_observed = "No measurement time available"
                    
    #                 # Set the peak max temperature with °C
    #                 record.peak_max_temp_diff = f"{max_temp_diff:.1f} °C"
    #         else:
    #             # Default values when no child lines exist
    #             record.peak_max_temp_diff = "0.0 °C"
    #             record.peak_max_temp_diff_observed = "No data"



    # graph_image_20mm = fields.Binary("Line Chart", compute="_compute_graph_image_20mm", store=True)




    # def generate_line_chart_20mm(self):
    #     import matplotlib.pyplot as plt
    #     import io
    #     import base64

    #     # Data preparation
    #     days = []  # Days for x-axis
    #     core_bottom = []
    #     core_middle = []
    #     core_top = []
    #     cover_bottom = []
    #     cover_middle = []
    #     cover_top = []
    #     ambient = []

    #     # Process data for alignment
    #     for line in self.child_lines_temprature:
    #         day = line.day
    #         if isinstance(line.bottom, list):  # Handle multiple entries for the same day
    #             for i in range(len(line.bottom)):
    #                 days.append(day)  # Same day for all entries
    #                 core_bottom.append(line.bottom[i])
    #                 core_middle.append(line.middle[i])
    #                 core_top.append(line.top[i])
    #                 cover_bottom.append(line.bottom1[i])
    #                 cover_middle.append(line.middle1[i])
    #                 cover_top.append(line.top1[i])
    #                 ambient.append(line.ambient[i])
    #         else:  # Single value per day
    #             days.append(day)
    #             core_bottom.append(line.bottom)
    #             core_middle.append(line.middle)
    #             core_top.append(line.top)
    #             cover_bottom.append(line.bottom1)
    #             cover_middle.append(line.middle1)
    #             cover_top.append(line.top1)
    #             ambient.append(line.ambient)

    #     # Create the plot
    #     plt.figure(figsize=(14, 8))

    #     # Create a list of x positions for each entry
    #     x_positions = list(range(len(days)))  # Create x positions based on the number of entries

    #     # Plot the temperature data using x positions for plotting
    #     plt.plot(x_positions, core_bottom, label='Core Area - Bottom', marker='o', color='blue')
    #     plt.plot(x_positions, core_middle, label='Core Area - Middle', marker='o', color='orange')
    #     plt.plot(x_positions, core_top, label='Core Area - Top', marker='o', color='green')
    #     plt.plot(x_positions, cover_bottom, label='Cover Area - Bottom', marker='o', color='red')
    #     plt.plot(x_positions, cover_middle, label='Cover Area - Middle', marker='o', color='purple')
    #     plt.plot(x_positions, cover_top, label='Cover Area - Top', marker='o', color='brown')
    #     plt.plot(x_positions, ambient, label='Ambient', linestyle='--', color='black')

    #     # Add vertical lines for each day
    #     for pos in x_positions:
    #         plt.axvline(x=pos, color='cyan', linestyle='-', linewidth=1)  # Draw vertical line

    #     # Add day labels inside the plot area for each position
    #     for i, day in enumerate(days):
    #         # Place the label slightly below the minimum y limit of the plot
    #         plt.text(i, plt.ylim()[0] - (0.05 * (plt.ylim()[1] - plt.ylim()[0])), 
    #                 day, color='cyan', rotation=90, fontsize=10,
    #                 verticalalignment='top', horizontalalignment='center')

    #     # Customize the graph
    #     plt.xlabel('Days', fontsize=14)
    #     plt.ylabel('Temperature (°C)', fontsize=14)
    #     plt.xticks(ticks=x_positions, labels=days, fontsize=10, rotation=45)  # Set x-ticks to days
    #     plt.yticks(fontsize=12)
    #     plt.legend(fontsize=12)
    #     plt.grid(True, which='both', linestyle='--', linewidth=0.5)

    #     # Save the plot to a binary field
    #     buf = io.BytesIO()
    #     plt.savefig(buf, format='png', bbox_inches='tight')
    #     buf.seek(0)
    #     chart_image20mm = base64.b64encode(buf.read()).decode('utf-8')
    #     buf.close()
    #     plt.close()  # Free up memory

    #     return chart_image20mm






    # @api.depends('child_lines_temprature')
    # def _compute_graph_image_20mm(self):
    #     try:
    #         for record in self:
    #             chart_image20mm = record.generate_line_chart_20mm()
    #             record.graph_image_20mm = chart_image20mm
    #     except:
    #         pass 


    # upload_file = fields.Binary(string="Upload Excel File")
    # file_name = fields.Char(string="Filename")



   

    # def action_import_excel(self):
    #     def safe_float(val):
    #         try:
    #             return float(val)
    #         except (ValueError, TypeError):
    #             return 0.0

    #     if not self.upload_file:
    #         raise UserError(_("Please upload an Excel file."))

    #     if not self.file_name.lower().endswith('.xlsx'):
    #         raise UserError(_("Only .xlsx files are supported."))

    #     try:
    #         file_data = base64.b64decode(self.upload_file)
    #         file_io = io.BytesIO(file_data)
    #         wb = openpyxl.load_workbook(file_io, data_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Could not read the Excel file: %s") % e)

    #     lines = []
    #     for row in sheet.iter_rows(min_row=2, values_only=True):
    #         if not any(row):
    #             continue

    #         lines.append((0, 4, {
    #             'measurement_time': str(row[1]).strip() if row[1] else '',
    #             'bottom': safe_float(row[2]),
    #             'middle': safe_float(row[3]),
    #             'top': safe_float(row[4]),
    #             'temp_diff': safe_float(row[5]),
    #             'bottom1': safe_float(row[6]),
    #             'middle1': safe_float(row[7]),
    #             'top1': safe_float(row[8]),
    #             'temp_diff1': safe_float(row[9]),
    #             'ambient': safe_float(row[10]),
    #         }))

    #     self.child_lines_temprature = lines

    introduction = fields.Text(string="Introduction")

    raft_thickness = fields.Char(string="Raft Thickness:")
    date_pouring = fields.Date(string="Date of Pouring:")
    date_termination = fields.Date(string="Date of Termination :")
    period_monitoring = fields.Char(string="Period of Monitoring: ")

    tempreture_upload1 = fields.Many2many(
        'ir.attachment',
        'tempreture_upload1_rel',
        'sample_id',
        'attachment_id',
        string='Upload Image',
        help='Attach multiple reports or PDFs to the sample',
    )


    temprerature_location = fields.Text(string="Temperature Monitoring location:")

    tempreture_upload2 = fields.Many2many(
        'ir.attachment',
        'tempreture_upload2_rel',
        'sample_id',
        'attachment_id',
        string='Upload Image',
        help='Attach multiple reports or PDFs to the sample',
    )

    position_thermpcouples = fields.Text(string="Positioning of the Thermocouples")

    result = fields.Text(string="Result")

    tempreture_upload3 = fields.Many2many(
        'ir.attachment',
        'tempreture_upload3_rel',
        'sample_id',
        'attachment_id',
        string='Upload Image',
        help='Attach multiple reports or PDFs to the sample',
    )



    temprature_line = fields.One2many('tempraturee.line','parent_id',string="Parameter", copy=True)

    show_thermocouple1_bottom = fields.Boolean(string="Show Thermocouple -01")
    show_thermocouple2_middle = fields.Boolean(string="Show Thermocouple -02")
    show_thermocouple3_top = fields.Boolean(string="Show Thermocouple -03")
    show_thermocouple4_stand1 = fields.Boolean(string="Show Stand By1 - 04")
    show_thermocouple5_stand2 = fields.Boolean(string="Show Stand By2 - 05")
    show_thermocouple6_bottom = fields.Boolean(string="Show Thermocouple -06")
    show_thermocouple7_middle = fields.Boolean(string="Show Thermocouple -07")
    show_thermocouple8_top = fields.Boolean(string="Show Thermocouple -08")
    show_thermocouple9_stand1 = fields.Boolean(string="Show Stand By1 - 09")
    show_thermocouple10_stand2 = fields.Boolean(string="Show Stand By2 - 10")
    show_thermocouple11_bottom = fields.Boolean(string="Show Thermocouple -11")
    show_thermocouple12_middle = fields.Boolean(string="Show Thermocouple -12")
    show_thermocouple13_top = fields.Boolean(string="Show Thermocouple -13")
    show_thermocouple14_stand1 = fields.Boolean(string="Show Stand By1 - 14")
    show_thermocouple15_stand2 = fields.Boolean(string="Show Stand By2 - 15")
    show_thermocouple16_bottom = fields.Boolean(string="Show Thermocouple -16")
    show_thermocouple17_middle = fields.Boolean(string="Show Thermocouple -17")
    show_thermocouple18_top = fields.Boolean(string="Show Thermocouple -18")
    show_thermocouple19_stand1 = fields.Boolean(string="Show Stand By1 - 19")
    show_thermocouple20_stand2 = fields.Boolean(string="Show Stand By2 - 20")
    show_ambient = fields.Boolean(string="Show Ambient")


    show_temperature_differential1 = fields.Boolean(string="Show Temperature Differential-01")
    show_temperature_differential2 = fields.Boolean(string="Show Temperature Differential-02")
    show_temperature_differential3 = fields.Boolean(string="Show Temperature Differential-03")
    show_temperature_differential4 = fields.Boolean(string="Show Temperature Differential-04")



    upload_file1 = fields.Binary(string="Upload Excel File")
    file_name = fields.Char(string="Filename")




    def action_import_excel_data(self):
        import base64
        import io
        import openpyxl
        from odoo.exceptions import UserError
        from odoo.tools.translate import _

        # ✅ Safe type conversions
        def safe_str(row, index):
            return str(row[index]).strip() if len(row) > index and row[index] else ''

        def safe_float(row, index):
            try:
                return float(row[index]) if len(row) > index and row[index] is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        def safe_int(row, index):
            try:
                return int(row[index]) if len(row) > index and row[index] is not None else 0
            except (ValueError, TypeError):
                return 0

        # ✅ Step 1: Validation
        if not self.upload_file1:
            raise UserError(_("Please upload an Excel file."))

        if not self.file_name or not isinstance(self.file_name, str) or not self.file_name.lower().endswith('.xlsx'):
            raise UserError(_("Only .xlsx files are supported."))

        # ✅ Step 2: Read Excel
        try:
            file_data = base64.b64decode(self.upload_file1)
            file_io = io.BytesIO(file_data)
            wb = openpyxl.load_workbook(file_io, data_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Could not read the Excel file: %s") % e)

        # ✅ Step 3: Row data import
        lines = []
        for row in sheet.iter_rows(min_row=4, values_only=True):  # Start from row 5
            if not any(row):
                continue

            lines.append((0, 4, {
                'days': safe_int(row, 0),
                'date': safe_str(row, 1),
                'time': safe_str(row, 2),
                'thermocouple1_bottom': safe_float(row, 3),
                'thermocouple2_middle': safe_float(row, 4),
                'thermocouple3_top': safe_float(row, 5),
                'thermocouple4_stand1': safe_str(row, 6),
                'thermocouple5_stand2': safe_str(row, 7),
                'thermal_gradient1': safe_str(row, 8),
                'thermocouple6_bottom': safe_float(row, 9),
                'thermocouple7_middle': safe_float(row, 10),
                'thermocouple8_top': safe_float(row, 11),
                'thermocouple9_stand1': safe_str(row, 12),
                'thermocouple10_stand2': safe_str(row, 13),
                'thermal_gradient2': safe_str(row, 14),
                'thermocouple11_bottom': safe_float(row, 15),
                'thermocouple12_middle': safe_float(row, 16),
                'thermocouple13_top': safe_float(row, 17),
                'thermocouple14_stand1': safe_str(row, 18),
                'thermocouple15_stand2': safe_str(row, 19),
                'thermal_gradient3': safe_str(row, 20),
                
                'thermocouple16_bottom': safe_float(row, 21),
                'thermocouple17_middle': safe_float(row, 22),
                'thermocouple18_top': safe_float(row, 23),
                'thermocouple19_stand1': safe_str(row, 24),
                'thermocouple20_stand2': safe_str(row, 25),
                'thermal_gradient4': safe_str(row, 26),
                'ambient1': safe_float(row, 27)
            }))

        # ✅ Step 4: Assign One2many
        self.temprature_line = lines


    def generate_line_chart_thermocouples(self):
        lines = self.temprature_line.sorted('days')
        x_values = list(range(1, len(lines) + 1))  # 1 pasun numbering

        # (field_name, label, show_boolean_field)
        tc_fields = [
            ('thermocouple1_bottom', 'Thermocouple -01', 'show_thermocouple1_bottom'),
            ('thermocouple2_middle', 'Thermocouple -02', 'show_thermocouple2_middle'),
            ('thermocouple3_top', 'Thermocouple -03', 'show_thermocouple3_top'),
            ('thermocouple4_stand1', 'Stand By1 - 04', 'show_thermocouple4_stand1'),
            ('thermocouple5_stand2', 'Stand By2 - 05', 'show_thermocouple5_stand2'),
            ('thermocouple6_bottom', 'Thermocouple -06', 'show_thermocouple6_bottom'),
            ('thermocouple7_middle', 'Thermocouple -07', 'show_thermocouple7_middle'),
            ('thermocouple8_top', 'Thermocouple -08', 'show_thermocouple8_top'),
            ('thermocouple9_stand1', 'Stand By1 - 09', 'show_thermocouple9_stand1'),
            ('thermocouple10_stand2', 'Stand By2 - 10', 'show_thermocouple10_stand2'),
            ('thermocouple11_bottom', 'Thermocouple -11', 'show_thermocouple11_bottom'),
            ('thermocouple12_middle', 'Thermocouple -12', 'show_thermocouple12_middle'),
            ('thermocouple13_top', 'Thermocouple -13', 'show_thermocouple13_top'),
            ('thermocouple14_stand1', 'Stand By1 - 14', 'show_thermocouple14_stand1'),
            ('thermocouple15_stand2', 'Stand By2 - 15', 'show_thermocouple15_stand2'),
            ('thermocouple16_bottom', 'Thermocouple -16', 'show_thermocouple16_bottom'),
            ('thermocouple17_middle', 'Thermocouple -17', 'show_thermocouple17_middle'),
            ('thermocouple18_top', 'Thermocouple -18', 'show_thermocouple18_top'),
            ('thermocouple19_stand1', 'Stand By1 - 19', 'show_thermocouple19_stand1'),
            ('thermocouple20_stand2', 'Stand By2 - 20', 'show_thermocouple20_stand2'),
        ]

        dark_colors = [
            '#1f77b4', '#2ca02c', '#d62728', '#9467bd', '#8c564b',
            '#e377c2', '#7f7f7f', '#bcbd22', '#17becf', '#ff7f0e',
            '#393b79', '#637939', '#8c6d31', '#843c39', '#7b4173',
            '#5254a3', '#6b6ecf', '#9c9ede', '#637939', '#8ca252'
        ]

        import matplotlib.pyplot as plt
        import io
        import base64

        plt.figure(figsize=(16, 9))
        color_idx = 0

        for field, label, show_field in tc_fields:
            if not getattr(self, show_field, False):
                continue

            values = []
            for rec in lines:
                val = getattr(rec, field, None)
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    values.append(None)

            color = dark_colors[color_idx % len(dark_colors)]
            color_idx += 1
            plt.plot(x_values, values, label=label, color=color)

        # ✅ Ambient line (always visible)
        ambient_values = []
        for rec in lines:
            val = getattr(rec, 'ambient1', None)
            try:
                ambient_values.append(float(val))
            except (ValueError, TypeError):
                ambient_values.append(None)

        plt.plot(
            x_values, ambient_values,
            label='Ambient Temperature',
            color='#7f2547',
            linestyle='-',
            linewidth=2
        )

        # Vertical day lines
        for x in x_values:
            plt.axvline(x=x, color='gray', linestyle='--', linewidth=0.3)

        # X-axis labels
        for i, day in enumerate(x_values):
            plt.text(day, plt.ylim()[0] - 1, str(day), color='black', rotation=90,
                    fontsize=9, verticalalignment='top', horizontalalignment='center')

        # Formatting
        plt.xlabel("Days", fontsize=14, fontweight='bold')
        plt.ylabel("Temperature (°C)", fontsize=14, fontweight='bold')
        plt.title("Thermocouple Temperature Graph", fontsize=16, fontweight='bold')
        plt.xticks(x_values, [str(x) for x in x_values], rotation=45)
        plt.grid(True, linestyle='--', linewidth=0.5)

        # Legend
        legend = plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
        for text in legend.get_texts():
            text.set_fontweight('bold')

        # Save to base64
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        chart_image = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()
        plt.close()

        return chart_image




    graph_image_tc = fields.Binary("Graph Image", compute="_compute_graph_image_tc",store=False)

    

    @api.depends('temprature_line')
    def _compute_graph_image_tc(self):
        try:
            for record in self:
                chart_image = record.generate_line_chart_thermocouples()
                record.graph_image_tc = chart_image
        except:
            pass 

    # @api.depends('temprature_line')
    # def _compute_graph_image_tc(self):
    #     for record in self:
    #         try:
    #             record.graph_image_tc = record.generate_line_chart_thermocouples()
    #         except:
    #             record.graph_image_tc = False




    graph_image_temp = fields.Binary(
        "Temperature Differential Graph",
        compute="_compute_graph_image_diff",
        store=True
    )

    @api.depends(
        'temprature_line.thermal_gradient1',
        'temprature_line.thermal_gradient2',
        'temprature_line.thermal_gradient3',
        'temprature_line.thermal_gradient4',
        'temprature_line.days'
    )
    def _compute_graph_image_diff(self):
        for record in self:
            try:
                chart_image_diff = record.generate_line_chart_tempreture_diff()
                record.graph_image_temp = chart_image_diff
            except Exception as e:
                _logger.error("Chart generation failed for record ID %s: %s", record.id, e)
                record.graph_image_temp = False  # Clear on failure

    def generate_line_chart_tempreture_diff(self):
        lines = self.temprature_line.sorted('days')
        x_values = list(range(1, len(lines) + 1))  # Start from 1

        temp_diff = [
            ('thermal_gradient1', 'Temperature Differential-01', 'show_temperature_differential1'),
            ('thermal_gradient2', 'Temperature Differential-02', 'show_temperature_differential2'),
            ('thermal_gradient3', 'Temperature Differential-03', 'show_temperature_differential3'),
            ('thermal_gradient4', 'Temperature Differential-04', 'show_temperature_differential4')
        ]

        dark_colors = ['#1f77b4', '#2ca02c', '#d62728', '#9467bd']

        plt.figure(figsize=(16, 9))
        color_idx = 0

        for field, label, show_field in temp_diff:
            if not getattr(self, show_field, False):
                continue

            values = []
            for rec in lines:
                val = getattr(rec, field, None)
                try:
                    values.append(float(val))
                except (ValueError, TypeError):
                    values.append(None)

            color = dark_colors[color_idx % len(dark_colors)]
            color_idx += 1
            plt.plot(x_values, values, label=label, color=color)

        
        # Vertical day lines
        for x in x_values:
            plt.axvline(x=x, color='gray', linestyle='--', linewidth=0.3)

        # X-axis labels
        for i, day in enumerate(x_values):
            plt.text(day, plt.ylim()[0] - 1, str(day), color='black', rotation=90,
                     fontsize=9, verticalalignment='top', horizontalalignment='center')

        # Labels and formatting
        plt.xlabel("Days", fontsize=14, fontweight='bold')
        plt.ylabel("Temperature (°C)", fontsize=14, fontweight='bold')
        plt.title("Thermocouple Temperature Graph", fontsize=16, fontweight='bold')
        plt.xticks(x_values, [str(x) for x in x_values], rotation=45)
        plt.grid(True, linestyle='--', linewidth=0.5)

        # Legend
        legend = plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize=9)
        for text in legend.get_texts():
            text.set_fontweight('bold')

        # Save to base64
        buf = io.BytesIO()
        plt.tight_layout()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        chart_image_diff = base64.b64encode(buf.read()).decode('utf-8')
        buf.close()
        plt.close()

        return chart_image_diff


   

















    def open_eln_page(self):
        # import wdb; wdb.set_trace()

        return {
                'view_mode': 'form',
                'res_model': "lerm.eln",
                'type': 'ir.actions.act_window',
                'target': 'current',
                'res_id': self.eln_ref.id,
                
            }           


    @api.model
    def create(self, vals):
        # import wdb;wdb.set_trace()
        record = super(EnviromentTemprature, self).create(vals)
        # record.get_all_fields()
        record.eln_ref.write({'model_id':record.id})
        return record







    @api.depends('eln_ref')
    def _compute_sample_parameters(self):
        # records = self.env['lerm.eln'].sudo().search([('id','=', record.eln_id.id)]).parameters_result
        # print("records",records)
        # self.sample_parameters = records
        for record in self:
            records = record.eln_ref.parameters_result.parameter.ids
            record.sample_parameters = records
            print("Records",records)



    def get_all_fields(self):
        record = self.env['enviroment.temprature'].browse(self.ids[0])
        field_values = {}
        for field_name, field in record._fields.items():
            field_value = record[field_name]
            field_values[field_name] = field_value

        return field_values




# class EnviromentTempratureLine(models.Model):
#     _name = "enviroment.tempraturee.line"
#     parent_id = fields.Many2one('enviroment.temprature',string="Parent Id")
   
#     sr_no = fields.Integer(string="Sr No.",readonly=True, copy=False, default=1)

#     day = fields.Char(string="Day")
#     date = fields.Date("Date")

   
#     measurement_time = fields.Char(string="Time")
#     bottom = fields.Float(string="bottom",digits=(12,1))
#     middle = fields.Float(string="Middle",digits=(12,1))
#     top = fields.Float(string="Top",digits=(12,1))
#     temp_diff = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_temp_diff", store=True)

     
#     bottom1 = fields.Float(string="bottom",digits=(12,1))
#     middle1 = fields.Float(string="Middle",digits=(12,1))
#     top1 = fields.Float(string="Top",digits=(12,1))
#     temp_diff1 = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_temp_diff1", store=True)

#     ambient = fields.Float(string="Ambient",digits=(12,1))


#     @api.depends('middle', 'top')
#     def _compute_temp_diff(self):
#         for record in self:
#             record.temp_diff = record.middle - record.top

#     @api.depends('middle1', 'top1')
#     def _compute_temp_diff1(self):
#         for record in self:
#             record.temp_diff1 = record.top1 - record.middle1




    

#     @api.model
#     def create(self, vals):
#         # Set the serial_no based on the existing records for the same parent
#         if vals.get('parent_id'):
#             existing_records = self.search([('parent_id', '=', vals['parent_id'])])
#             if existing_records:
#                 max_serial_no = max(existing_records.mapped('sr_no'))
#                 vals['sr_no'] = max_serial_no + 1

#         return super(EnviromentTempratureLine, self).create(vals)

#     def _reorder_serial_numbers(self):
#         # Reorder the serial numbers based on the positions of the records in child_lines
#         records = self.sorted('id')
#         for index, record in enumerate(records):
#             record.sr_no = index + 1




class TempratureLine(models.Model):
    _name = "tempraturee.line"
    parent_id = fields.Many2one('enviroment.temprature',string="Parent Id")
   
    sr_no = fields.Integer(string="Sr No.",readonly=True, copy=False, default=1)

    days = fields.Integer("Days")

    date = fields.Char("Date")


    # show_thermocouple1_bottom = fields.Boolean(
    #     string="Show Thermocouple -01 bottom",
    #     compute="_compute_show_thermocouple1_bottom",
    #     store=True,
    # )


    # @api.depends('parent_id.show_thermocouple1_bottom')
    # def _compute_show_thermocouple1_bottom(self):
    #     for rec in self:
    #         rec.show_thermocouple1_bottom = rec.parent_id.show_thermocouple1_bottom
   
    time = fields.Char(string="Time")

    thermocouple1_bottom = fields.Float(string="Thermocouple -01 bottom",digits=(12,1))
    thermocouple2_middle = fields.Float(string="Thermocouple -02 Middle",digits=(12,1))
    thermocouple3_top = fields.Float(string="Thermocouple -03 Top",digits=(12,1))
    thermocouple4_stand1 = fields.Char(string="Thermocouple -04 Stand By1")
    thermocouple5_stand2 = fields.Char(string="Thermocouple -05 Stand By1")
    thermal_gradient1 = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_thermal_gradient1", store=True)

    @api.depends('thermocouple2_middle', 'thermocouple3_top')
    def _compute_thermal_gradient1(self):
        for record in self:
            record.thermal_gradient1 = record.thermocouple2_middle - record.thermocouple3_top

     
    thermocouple6_bottom = fields.Float(string="Thermocouple -06 bottom",digits=(12,1))
    thermocouple7_middle = fields.Float(string="Thermocouple -07 Middle",digits=(12,1))
    thermocouple8_top = fields.Float(string="Thermocouple -08 Top",digits=(12,1))
    thermocouple9_stand1 = fields.Char(string="Thermocouple -09 Stand By1")
    thermocouple10_stand2 = fields.Char(string="Thermocouple -10 Stand By1")
    thermal_gradient2 = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_thermal_gradient2", store=True)

    @api.depends('thermocouple7_middle', 'thermocouple8_top')
    def _compute_thermal_gradient2(self):
        for record in self:
            record.thermal_gradient2 = record.thermocouple7_middle - record.thermocouple8_top

    thermocouple11_bottom = fields.Float(string="Thermocouple -11 bottom",digits=(12,1))
    thermocouple12_middle = fields.Float(string="Thermocouple -12 Middle",digits=(12,1))
    thermocouple13_top = fields.Float(string="Thermocouple -13 Top",digits=(12,1))
    thermocouple14_stand1 = fields.Char(string="Thermocouple -14 Stand By1")
    thermocouple15_stand2 = fields.Char(string="Thermocouple -15 Stand By1")
    thermal_gradient3 = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_thermal_gradient3", store=True)

    @api.depends('thermocouple12_middle', 'thermocouple13_top')
    def _compute_thermal_gradient3(self):
        for record in self:
            record.thermal_gradient3 = record.thermocouple12_middle - record.thermocouple13_top

    thermocouple16_bottom = fields.Float(string="Thermocouple -16 bottom",digits=(12,1))
    thermocouple17_middle = fields.Float(string="Thermocouple -17 Middle",digits=(12,1))
    thermocouple18_top = fields.Float(string="Thermocouple -18 Top",digits=(12,1))
    thermocouple19_stand1 = fields.Char(string="Thermocouple -19 Stand By1")
    thermocouple20_stand2 = fields.Char(string="Thermocouple -20 Stand By1")
    thermal_gradient4 = fields.Float(string="Thermal Gradient",digits=(12,1),compute="_compute_thermal_gradient4", store=True)

    @api.depends('thermocouple17_middle', 'thermocouple18_top')
    def _compute_thermal_gradient4(self):
        for record in self:
            record.thermal_gradient4 = record.thermocouple17_middle - record.thermocouple18_top

    ambient1 = fields.Float(string="Ambient",digits=(12,1))



   




    

    @api.model
    def create(self, vals):
        # Set the serial_no based on the existing records for the same parent
        if vals.get('parent_id'):
            existing_records = self.search([('parent_id', '=', vals['parent_id'])])
            if existing_records:
                max_serial_no = max(existing_records.mapped('sr_no'))
                vals['sr_no'] = max_serial_no + 1

        return super(TempratureLine, self).create(vals)

    def _reorder_serial_numbers(self):
        # Reorder the serial numbers based on the positions of the records in child_lines
        records = self.sorted('id')
        for index, record in enumerate(records):
            record.sr_no = index + 1


   